"""Golden solution: Cutover_Plan.xlsx, formula-driven throughout.

Typed cells are limited to source data on the Reference sheet and the three
planning decisions per site (team, start week, wave). Everything else computes.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import cutover as c
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL
OUT="/home/azureuser/geranium_tasks/task2_env/golden_solution/Cutover_Plan.xlsx"
SLATE,PALE,AMBER,GREY,RED="2E4057","DCE3EB","FFF2CC","EDEDED","F2DCDB"
HDR=Font(bold=True,color="FFFFFF",size=9.5); THIN=Side(style="thin",color="BFBFBF")
BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
IN_F=Font(color="0033CC",size=9.5); CA_F=Font(color="000000",size=9.5)
D0='yyyy-mm-dd'
sch=c.solve(); wv=c.waves(sch); order=sorted(sch,key=lambda x:(sch[x]['start'],x))
BL=sorted(w for w in c.BLACKOUT if w<=c.LAST_WEEK)
LW=c.LAST_WEEK

def hdr(ws,r,hs,w=None):
    for i,h in enumerate(hs,1):
        x=ws.cell(r,i,h); x.font,x.border=HDR,BOX; x.fill=PatternFill("solid",fgColor=SLATE)
        x.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    ws.row_dimensions[r].height=30
    for i,v in (w or {}).items(): ws.column_dimensions[CL(i)].width=v
def put(ws,r,vals,fmts=None,inputs=(),bold=False,fill=None):
    for i,v in enumerate(vals,1):
        x=ws.cell(r,i,v); x.border=BOX
        x.font=Font(bold=True,size=9.5) if bold else (IN_F if i in inputs else CA_F)
        if fill: x.fill=PatternFill("solid",fgColor=fill)
        if fmts and fmts.get(i): x.number_format=fmts[i]
        x.alignment=Alignment(wrap_text=True,vertical="top")
    return r+1
def title(ws,t,sub=None):
    ws.cell(1,1,t).font=Font(bold=True,size=13,color=SLATE)
    if sub: ws.cell(2,1,sub).font=Font(italic=True,size=9,color="595959")
    ws.sheet_view.showGridLines=False
    return 4

wb=Workbook()
# ============================================================ Reference ====
ws=wb.active; ws.title="Reference"
title(ws,"Reference data",
      "Transcribed from the site profile, the programme brief and the vendor notice. "
      "Blue cells are typed source data; everything on the other sheets computes from here.")
for i,w in {1:38,2:16,3:16,4:12,5:20,6:14,7:10,8:14,9:8}.items(): ws.column_dimensions[CL(i)].width=w
put(ws,4,["Programme week 1 commences",c.PROGRAMME_START],fmts={2:D0},inputs=(2,))
put(ws,5,["Deployment teams",c.TEAMS],inputs=(2,))
put(ws,6,["Validation weeks, regulated site",c.VALIDATION_WEEKS],inputs=(2,))
put(ws,7,["Legacy support ends",c.LEGACY_SUPPORT_ENDS],fmts={2:D0},inputs=(2,))
put(ws,8,["Support ends, programme week","=INT((B7-B4)/7)+1"])
hdr(ws,10,["Code","Site","Country","Users","Master data owned","Depends on","Weeks",
           "Shutdown week","GMP"])
r=11
for code,s in c.SITE.items():
    put(ws,r,[code,s["name"],s["country"],s["users"],s["master"] or "",
              ", ".join(s["deps"]) or "",s["weeks"],
              s["shutdown"][0] if s["shutdown"] else "","Yes" if s["gmp"] else "No"],
        inputs=tuple(range(1,10))); r+=1
SITE_TOP,SITE_BOT=11,r-1
ws.cell(26,1,"Financial blackout weeks").font=Font(bold=True,size=9.5)
for i,w in enumerate(BL): put(ws,26,[None]*0) if False else ws.cell(26,2+i,w)
for i in range(len(BL)):
    x=ws.cell(26,2+i); x.font=IN_F; x.border=BOX
BLROW=26; BLLAST=CL(1+len(BL))
REF=f"Reference!$A${SITE_TOP}:$A${SITE_BOT}"
def lk(col,codecell): return f"INDEX(Reference!${col}${SITE_TOP}:${col}${SITE_BOT},MATCH({codecell},{REF},0))"

# ================================================================= Plan ====
ws=wb.create_sheet("Plan")
r=title(ws,f"{c.CLIENT} ERP cutover sequence",
        "Blue cells are the three planning decisions per site: wave, team and start week. "
        "Every other column computes from the Reference sheet.")
hdr(ws,r,["Wave","Code","Site","Country","Users","Team","Start week","Week commencing",
          "Weeks","End week","Live week","Live date","Binding constraint"],
    {1:6,2:7,3:13,4:15,5:8,6:6,7:10,8:15,9:7,10:10,11:10,12:14,13:50})
r+=1; PTOP=r
for code in order:
    s=c.SITE[code]; p=sch[code]; note=[]
    if s["master"]: note.append(f"Owns {s['master'].lower()}; {sum(1 for x in c.SITE.values() if code in x['deps'])} sites depend on it")
    if s["gmp"]: note.append("Validation runs after cutover, so it must start early despite being the smallest site")
    if s["shutdown"]: note.append("Can only start in its annual shutdown window")
    if s["deps"] and not note: note.append("Follows "+", ".join(s["deps"]))
    put(ws,r,[wv[code],code,f"={lk('B',f'$B{r}')}",f"={lk('C',f'$B{r}')}",f"={lk('D',f'$B{r}')}",
              p["team"],p["start"],f"=Reference!$B$4+($G{r}-1)*7",f"={lk('G',f'$B{r}')}",
              f"=$G{r}+$I{r}-1",
              f"=$J{r}+1+IF({lk('I',f'$B{r}')}=\"Yes\",Reference!$B$6,0)",
              f"=Reference!$B$4+($K{r}-1)*7","; ".join(note)],
        fmts={5:'#,##0',8:D0,12:D0},inputs=(1,2,6,7,13)); r+=1
PBOT=r-1
put(ws,r,["","TOTAL","","",f"=SUM(E{PTOP}:E{PBOT})","","","",f"=SUM(I{PTOP}:I{PBOT})","",
          f"=MAX(K{PTOP}:K{PBOT})",f"=Reference!$B$4+(K{r}-1)*7",
          "Last go-live across the programme"],
    fmts={5:'#,##0',12:D0},bold=True,fill=GREY); TOTROW=r; r+=1
put(ws,r,["","Support deadline","","",f"=Reference!$B$8","","","","","",f"=Reference!$B$8",
          f"=Reference!$B$7","Programme week in which legacy support ends"],fmts={12:D0}); r+=1
put(ws,r,["","Float to the deadline","","",f"=Reference!$B$8-K{TOTROW}","","","","","","","",
          "Weeks between the last go-live and the end of support"],bold=True,fill=AMBER)

# ============================================================= Timeline ====
ws=wb.create_sheet("Timeline")
r=title(ws,"Team loading by programme week",
        "Every cell computes from the Plan sheet. Row 3 flags the financial blackout weeks.")
hdr(ws,4,["Code"]+[""]*LW,{1:8,**{i:5 for i in range(2,LW+2)}})
ws.cell(3,1,"Blackout").font=Font(bold=True,size=8.5)
for w in range(1,LW+1):
    ws.cell(4,w+1,w if w==1 else f"={CL(w)}4+1").font=HDR
    ws.cell(4,w+1).fill=PatternFill("solid",fgColor=SLATE); ws.cell(4,w+1).border=BOX
    ws.cell(4,w+1).alignment=Alignment(horizontal="center")
    b=ws.cell(3,w+1,f"=IF(COUNTIF(Reference!$B${BLROW}:${BLLAST}${BLROW},{CL(w+1)}$4)>0,1,0)")
    b.font=Font(size=8); b.border=BOX; b.alignment=Alignment(horizontal="center")
r=5
for i,code in enumerate(order):
    prow=PTOP+i
    row=[f"=Plan!$B{prow}"]+[f"=IF(AND(Plan!$G{prow}<={CL(w+1)}$4,Plan!$J{prow}>={CL(w+1)}$4),Plan!$B{prow},\"\")"
                             for w in range(1,LW+1)]
    put(ws,r,row); r+=1
TLTOP,TLBOT=5,r-1
put(ws,r,["Sites live"]+[f"=SUMPRODUCT((Plan!$G${PTOP}:$G${PBOT}<={CL(w+1)}$4)*(Plan!$J${PTOP}:$J${PBOT}>={CL(w+1)}$4))" for w in range(1,LW+1)],
    bold=True,fill=GREY); LOADROW=r; r+=2
for t in range(1,c.TEAMS+1):
    put(ws,r,[f"Team {t}"]+[f"=SUMPRODUCT((Plan!$F${PTOP}:$F${PBOT}={t})*(Plan!$G${PTOP}:$G${PBOT}<={CL(w+1)}$4)"
                            f"*(Plan!$J${PTOP}:$J${PBOT}>={CL(w+1)}$4))" for w in range(1,LW+1)]); r+=1
r+=1
put(ws,r,["Peak concurrent",f"=MAX($B${LOADROW}:${CL(LW+1)}${LOADROW})","Teams available",
          "=Reference!$B$5"],bold=True,fill=AMBER)
put(ws,r+1,["Cutovers in a blackout week",
            f"=SUMPRODUCT($B$3:${CL(LW+1)}$3,$B${LOADROW}:${CL(LW+1)}${LOADROW})",
            "Must be zero"],bold=True,fill=AMBER)
BLHIT=r+1

# =============================================================== Checks ====
ws=wb.create_sheet("Checks")
r=title(ws,"Constraint verification","Every result computes from the Plan and Timeline sheets.")
hdr(ws,r,["Constraint","Test","Result","Detail"],{1:30,2:46,3:12,4:60}); r+=1
def prow(code): return PTOP+order.index(code)
def trow(code): return TLTOP+order.index(code)
for code in order:
    for d in c.SITE[code]["deps"]:
        r=put(ws,r,[f"Master data, {code}",f"{code} starts after {d} completes",
                    f"=IF(Plan!G{prow(code)}>Plan!J{prow(d)},\"Pass\",\"FAIL\")",
                    f"{code} consumes {c.SITE[d]['master'].lower()} owned by {d}"])
for code in order:
    s=c.SITE[code]
    if s["shutdown"]:
        r=put(ws,r,[f"Shutdown window, {code}",f"{code} starts in its stated shutdown week",
                    f"=IF(Plan!G{prow(code)}={lk('H',f'Plan!$B{prow(code)}')},\"Pass\",\"FAIL\")",
                    "Continuous process plant; the turnaround date is fixed a year ahead"])
    if s["gmp"]:
        r=put(ws,r,[f"Validation, {code}","Live week is the week after cutover plus the validation period",
                    f"=IF(Plan!K{prow(code)}=Plan!J{prow(code)}+1+Reference!$B$6,\"Pass\",\"FAIL\")",
                    "The site is not live until validation completes"])
for code in order:
    r=put(ws,r,[f"Blackouts, {code}","No blackout week falls inside the cutover",
                f"=IF(SUMPRODUCT((Reference!$B${BLROW}:${BLLAST}${BLROW}>=Plan!$G{prow(code)})"
                f"*(Reference!$B${BLROW}:${BLLAST}${BLROW}<=Plan!$J{prow(code)}))=0,\"Pass\",\"FAIL\")",
                ""])
r=put(ws,r,["Team capacity","No week requires more teams than are available",
            f"=IF(MAX(Timeline!$B${LOADROW}:${CL(LW+1)}${LOADROW})<=Reference!$B$5,\"Pass\",\"FAIL\")",
            "Peak concurrency against the deployment teams available"])
r=put(ws,r,["Deadline","Every site live on or before the support deadline",
            f"=IF(MAX(Plan!K{PTOP}:K{PBOT})<=Reference!$B$8,\"Pass\",\"FAIL\")",
            "Cork's live date is the end of validation, not the end of cutover"],bold=True,fill=AMBER)
r=put(ws,r,["Durations","Team weeks consumed match the site profile",
            f"=IF(SUM(Plan!I{PTOP}:I{PBOT})=SUM(Reference!$G${SITE_TOP}:$G${SITE_BOT}),\"Pass\",\"FAIL\")",
            "No duration has been shortened to make the plan fit"])
r=put(ws,r,["All checks","Count of failures across the tests above",
            f"=COUNTIF(C5:C{r-1},\"FAIL\")","Must be zero"],bold=True,fill=AMBER)

# =========================================================== Sequencing ====
ws=wb.create_sheet("Sequencing")
r=title(ws,"Why the sequence is what it is")
ws.column_dimensions['A'].width=126
NDEP=sum(1 for x in c.SITE.values() if 'BRD' in x['deps'])
LAST=max(v['live'] for v in sch.values())
for t,b in [
 ("The order is set by freedom of movement, not by size",
  f"{NDEP} of the fourteen sites cannot start until Bridgnorth completes, because Bridgnorth owns "
  f"material master. Bridgnorth therefore leads regardless of anything else. Halesworth and Kinloss "
  f"own the other two master domains and run alongside it, so all three teams are committed in wave "
  f"one and the dependency chain is released in five weeks."),
 ("Charleston is the largest site and does not go first",
  f"Charleston has {c.SITE['CHA']['users']} users and is the obvious wave one candidate on any "
  f"size-ordered plan. It depends on both Bridgnorth and Halesworth and cannot start until week 10. "
  f"Sequencing by size is the error that sank the first internal attempt."),
 ("Cork is the smallest site and goes second",
  f"Cork has {c.SITE['COR']['users']} users and would be last on any plan ordered by size or by "
  f"revenue. It is GMP regulated and needs {c.VALIDATION_WEEKS} weeks of computer system validation "
  f"after the technical cutover before it can transact. Its cutover takes {c.SITE['COR']['weeks']} "
  f"weeks, so it must begin no later than programme week "
  f"{LW-c.SITE['COR']['weeks']-c.VALIDATION_WEEKS} to be live before support ends. Placing it early "
  f"removes the exposure entirely and costs nothing, because it is small enough to run beside a "
  f"larger site."),
 ("Sarnia and Mobile are fixed points, not choices",
  f"Both run continuous plant and can only begin in their annual maintenance shutdown. Sarnia's "
  f"window is programme week {c.SITE['SAR']['shutdown'][0]} and Mobile's is week "
  f"{c.SITE['MOB']['shutdown'][0]}. The rest of the plan is built around them rather than the "
  f"reverse, because those dates are set by turnaround contractors and will not move."),
 ("The blackouts remove a quarter of the calendar",
  f"Eleven of the {LW} programme weeks are unavailable: the week containing each quarter end and the "
  f"week before it, and the first three weeks of January. They fall in clusters, so the usable "
  f"calendar is four blocks rather than a continuous run. An evenly spread plan lands in them, which "
  f"is what happened to the second internal attempt."),
 ("Where the float is, and what it does not protect",
  f"The last go-live is programme week {LAST} against a deadline of {LW}, leaving {LW-LAST} weeks. "
  f"That float sits at the end of the programme and protects the smallest sites. It does not protect "
  f"Cork, whose validation is complete long before then, and it does not protect Sarnia or Mobile, "
  f"whose windows occur once a year."),
 ("What would break it",
  f"A slip on Bridgnorth moves {NDEP} sites. A slip that causes Sarnia or Mobile to miss its shutdown "
  f"window moves that site by a full year, which is past the deadline and past the point at which the "
  f"vendor will answer a support call. Those three sites carry the programme; the rest have room."),
]:
    p=ws.cell(r,1,t); p.font=Font(bold=True,size=10.5,color=SLATE); r+=1
    q=ws.cell(r,1,b); q.alignment=Alignment(wrap_text=True,vertical="top")
    ws.row_dimensions[r].height=max(15,12.4*(len(b)//122+1)); r+=2
wb.save(OUT)
# ---- census ----
from openpyxl import load_workbook
w2=load_workbook(OUT)
tot_t=tot_f=0
for name in ("Plan","Timeline","Checks"):
    s=w2[name]; t=f=0
    for row in s.iter_rows():
        for cl_ in row:
            v=cl_.value
            if v is None or v=="": continue
            if isinstance(v,str) and v.startswith("="): f+=1
            else: t+=1
    tot_t+=t; tot_f+=f
    print(f"  {name:10} typed {t:4}  formula {f:4}  typed share {t/(t+f):6.1%}")
print(f"  {'TOTAL':10} typed {tot_t:4}  formula {tot_f:4}  typed share {tot_t/(tot_t+tot_f):6.1%}  (threshold 80%)")
